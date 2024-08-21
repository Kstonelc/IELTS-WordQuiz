# -*- coding: utf-8 -*-
from flask import Blueprint, jsonify, request, send_file
from datetime import date, datetime
from urllib.parse import quote
import openpyxl
import os

words_api = Blueprint('words', __name__)


def analyze_file(file_path, iterating, first_word):
    print("当前要分析哪些轮:", file_path, iterating)
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    anasisly_col_index = []
    total_words_count = 0
    total_correct_words_count = 0
    current_total_words_count = 0
    current_correct_words_count = 0
    is_record_current = False
    # 获取需要分析的列
    for row in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            if str(cell.value).strip() in iterating:
                anasisly_col_index.append(cell.column)
    for col_index in anasisly_col_index:
        for row in sheet.iter_rows(min_row=4, min_col=col_index, max_col=col_index + 1):
            current_row = row[0].row
            correct_word = sheet.cell(row=current_row, column=4).value.strip()
            if not correct_word:
                continue
            row_data = []
            for cell in row:
                if cell.value is not None:
                    row_data.append(cell.value.strip())
            # 听写单词 和 订正
            if row_data:
                total_words_count += 1

                if row_data[0] is None:
                    is_record_current = False
                    break

                if correct_word == first_word:
                    is_record_current = True

                if is_record_current:
                    current_total_words_count += 1
                    if row_data[0] == correct_word:
                        current_correct_words_count += 1

                if row_data[0] == correct_word:
                    row[-2].fill = openpyxl.styles.PatternFill(
                                                start_color='FFD7D7', end_color='FFD7D7', fill_type='solid')
                    total_correct_words_count += 1
                else:
                    row[-2].fill = openpyxl.styles.PatternFill(
                                                    start_color='FF0000', end_color='FF0000', fill_type='solid')
                if len(row_data) == 2:
                    if row_data[1] == correct_word:
                        row[-1].fill = openpyxl.styles.PatternFill(
                            start_color='FFD7D7', end_color='FFD7D7', fill_type='solid')
                    else:
                        row[-1].fill = openpyxl.styles.PatternFill(
                            start_color='FF0000', end_color='FF0000', fill_type='solid')
        print("本次分析正确的单词个数:", current_correct_words_count, "\n本次分析总单词个数:", current_total_words_count)
        print("全部正确单词个数", total_correct_words_count, "\n全部单词个数:", total_words_count)
        sheet.cell(row=3, column=col_index).value = str(
            round(current_correct_words_count / current_total_words_count * 100, 2)) + "%"
        sheet.cell(row=2, column=col_index).value = str(
            round(total_correct_words_count / total_words_count * 100, 2)) + "%"
    current_formatted_date = datetime.now().strftime("%Y-%m-%d-%H-%M")
    file_path_without_extension = os.path.splitext(file_path)[0]
    full_file_name = f"{file_path_without_extension}-{current_formatted_date}.xlsx"
    wb.save(full_file_name)
    return full_file_name


@words_api.route('/words-analysis', methods=['POST'])
def words_analysis():
    data = request.get_json()
    file_name = data['file_name']
    iterating = data['iterating']
    first_word = data['first_word']
    file_path = os.path.join('./upload_file', file_name)

    try:
        result_file_path = analyze_file(file_path, iterating, first_word)
        result_file_name = os.path.basename(result_file_path)
        response = send_file(result_file_path,
                             as_attachment=True,
                             download_name=result_file_name,
                             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response.headers["content-disposition"] = f"attachment; filename*=utf-8''{quote(result_file_name)}"
        return response
    except Exception as e:
        print(e)
        response = {}
        response['ok'] = False
        response['data'] = "解析失败: " + str(e)
        return response


@words_api.route('/upload-file', methods=['POST'])
def upload_file():
    response = {}

    # 检查请求中是否包含文件
    if 'file' not in request.files:
        response['ok'] = False
        response['data'] = "没有文件"
        return jsonify(response), 200

    try:
        file = request.files['file']
        print(file)
        if file:
            filename = file.filename
            file_path = os.path.join('./upload_file', filename)
            file.save(file_path)
        response['ok'] = True
        response['data'] = "上传成功"
    except Exception as e:
        response['ok'] = False
        response['data'] = "上传失败: " + str(e)

    return jsonify(response), 200
